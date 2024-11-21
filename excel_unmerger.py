import os
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def list_excel_files():
    files = [f for f in os.listdir() if f.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')) and not f.startswith('unmerged_') and not f.startswith('~$')]
    return files

def copy_style(source_cell, target_cell):
    target_cell.font = source_cell.font.copy()
    target_cell.border = source_cell.border.copy()
    target_cell.fill = source_cell.fill.copy()
    target_cell.number_format = source_cell.number_format
    target_cell.protection = source_cell.protection.copy()
    target_cell.alignment = source_cell.alignment.copy()

def unmerge_cells(file_path, ignore_rows=0, ignore_cols=0):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        merged_cells = list(ws.merged_cells)
        for merged_cell in merged_cells:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
            if min_row > ignore_rows and min_col > ignore_cols:
                top_left_cell = ws.cell(row=min_row, column=min_col)
                top_left_value = top_left_cell.value
                ws.unmerge_cells(str(merged_cell))
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = ws.cell(row=row, column=col, value=top_left_value)
                        copy_style(top_left_cell, cell)
    new_file_path = f"unmerged_{os.path.basename(file_path)}"
    wb.save(new_file_path)
    print(f"Saved unmerged file as {new_file_path}")

def main():
    while True:
        files = list_excel_files()
        if not files:
            print("No Excel files to process.")
            break
        print("Excel files found:")
        for i, file in enumerate(files):
            print(f"{i + 1}. {file}")
        
        input("Press Enter to confirm the files to be processed...")

        ignore_rows = input("Enter the number of rows to ignore (leave empty to process entire file): ")
        ignore_cols = input("Enter the number of columns to ignore (leave empty to process entire file): ")

        ignore_rows = int(ignore_rows) if ignore_rows.isdigit() else 0
        ignore_cols = int(ignore_cols) if ignore_cols.isdigit() else 0

        for file in files:
            print(f"Processing file: {file}")
            try:
                unmerge_cells(file, ignore_rows, ignore_cols)
            except Exception as e:
                print(f"Failed to process {file}: {e}")
        
        input("Press Enter to process the next file or close the program to exit...")

if __name__ == "__main__":
    main()